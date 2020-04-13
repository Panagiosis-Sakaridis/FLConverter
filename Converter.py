import os
import tkinter as tk
from tkinter import filedialog

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl import Workbook

script_path = os.path.realpath(__file__)
script_path = script_path[0:script_path.rindex('\\')]


# noinspection DuplicatedCode
class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)

        # Global variables init ----------------------------------------------------------------------------------------
        self.in_filename = "coordinates.xlsx"
        self.out_filename = "coordinates_out.xlsx"
        self.driver_state = "stopped"
        self.driver = type(webdriver.firefox)
        self.start_driver()
        self.workingState = 0

        self.first_row = 2
        self.first_col = 2

        # Window init --------------------------------------------------------------------------------------------------
        self.master = master
        master.resizable(False, False)
        self.grid()

        # Window frames init -------------------------------------------------------------------------------------------
        self.file_frame = tk.Frame(self.master, relief="sunken", bg="grey")
        self.main_frame = tk.Frame(self.master)
        self.exit_frame = tk.Frame(self.master)

        # Variable strings init ----------------------------------------------------------------------------------------
        self.in_file_text = tk.StringVar()
        self.out_file_text = tk.StringVar()
        self.progress_text = tk.StringVar()
        self.option = tk.StringVar()
        self.option.set("web")

        # Label init ---------------------------------------------------------------------------------------------------
        self.file_title_label = tk.Label(self.main_frame,   text="Use these files:",          bg="grey"                )
        self.tip_label_in = tk.Label(self.file_frame,       text="Input file:   ",            bg="grey"                )
        self.tip_label_out = tk.Label(self.file_frame,      text="Output file:  ",            bg="grey"                )
        self.in_file_label = tk.Label(self.file_frame,      textvariable=self.in_file_text,   bg="grey"                )
        self.out_file_label = tk.Label(self.file_frame,     textvariable=self.out_file_text,  bg="grey"                )
        self.select_col_label = tk.Label(self.file_frame,   text="First cell: Col: ",         bg="grey"                )
        self.select_row_label = tk.Label(self.file_frame,   text="               Row: ",      bg="grey"                )
        self.web_title_label = tk.Label(self.main_frame,    text="Convert Using: "                                     )
        self.fl_label = tk.Label(self.main_frame,           text="     Φ,Λ     "                                       )
        self.hatt_label = tk.Label(self.main_frame,         text="        hatt        "                                )
        self.egsa_label = tk.Label(self.main_frame,         text="     egsa     "                                      )
        self.progress_label = tk.Label(self.main_frame, textvariable=self.progress_text,  fg="green", padx=10, pady=10 )
        self.ofl_title_label = tk.Label(self.main_frame,    text="Offline:"                                            )

        # Button init --------------------------------------------------------------------------------------------------
        self.select_file_in_button = tk.Button(self.file_frame,     bg="grey")
        self.open_file_in_button = tk.Button(self.file_frame,       bg="grey")
        self.select_file_out_button = tk.Button(self.file_frame,    bg="grey")
        self.open_file_out_button = tk.Button(self.file_frame,      bg="grey")
        self.fl_to_hatt_button = tk.Button(self.main_frame)
        self.hatt_to_egsa_button = tk.Button(self.main_frame)
        self.quit_button = tk.Button(self.exit_frame, text="QUIT",  fg="red")

        # Entries init -------------------------------------------------------------------------------------------------
        self.row_entry = tk.Entry(self.file_frame)
        self.row_entry.insert(0, str(self.first_row))
        self.col_entry = tk.Entry(self.file_frame)
        self.col_entry.insert(0, str(self.first_col))

        # Radio Buttons init -------------------------------------------------------------------------------------------
        self.web_rb = tk.Radiobutton(self.main_frame, text="Online", value="web", var=self.option)
        self.no_web_rb = tk.Radiobutton(self.main_frame, text="Offline", value="no_web", var=self.option)

        self.place_widgets()

    def destroy_self(self):
        if self.driver_state == "running":
            self.driver.quit()
        try:
            os.remove("geckodriver.log")
        except PermissionError:
            print("Exception raised: PermissionError. File: geckodriver.log")
        except FileNotFoundError:
            print("Exception raised: FileNotFoundError. File: geckodriver.log")

        self.master.destroy()

    def place_widgets(self):

        # File frame widgets -------------------------------------------------------------------------------------------
        self.tip_label_in.grid(                                         column=0, row=0, sticky='w'                    )
        self.in_file_text.set(self.in_filename)
        self.in_file_label.grid(                                        column=1, row=0, sticky='w'                    )

        self.select_file_in_button["text"] = "Change input file...  "
        self.select_file_in_button["command"] = self.change_file_in
        self.select_file_in_button.grid(                                column=2, row=0,            padx=5, pady=5     )

        self.open_file_in_button["text"] = "Open input file.  "
        self.open_file_in_button["command"] = self.open_file_in_pros
        self.open_file_in_button.grid(                                  column=3, row=0,            padx=10            )

        self.tip_label_out.grid(                                        column=0, row=1, sticky='w'                    )
        self.out_file_text.set(self.out_filename)
        self.out_file_label.grid(                                       column=1, row=1, sticky='w'                    )

        self.select_file_out_button["text"] = "Change output file..."
        self.select_file_out_button["command"] = self.change_file_out
#       self.select_file_out_button.grid(                               column=2, row=1,            padx=5             )

        self.open_file_out_button["text"] = "Open output file."
        self.open_file_out_button["command"] = self.open_file_out_pros
        self.open_file_out_button.grid(                                 column=3, row=1,            padx=5, pady=5     )
        self.select_col_label.grid(                                     column=0, row=2,                               )
        self.col_entry.grid(                                            column=1, row=2                                )
        self.select_row_label.grid(                                     column=0, row=3,                               )
        self.row_entry.grid(                                            column=1, row=3                                )

        self.file_frame.grid(                                                     row=0                                )

        # Main widgets -------------------------------------------------------------------------------------------------
        self.web_title_label.grid(                                      column=0, row=0,                               )
        self.web_rb.grid(                                               column=2, row=0                                )
        self.no_web_rb.grid(                                            column=3, row=0,            padx=10, pady=10   )

        self.fl_label.grid(                                             column=0, row=1                                )

        self.fl_to_hatt_button["text"] = "     -->     "
        self.fl_to_hatt_button["command"] = self.start_fl_to_hatt
        self.fl_to_hatt_button.grid(                                    column=1, row=1                                )

        self.hatt_label.grid(                                           column=2, row=1                                )

        self.hatt_to_egsa_button["text"] = "     -->     "
        self.hatt_to_egsa_button["command"] = self.start_hatt_to_egsa
        self.hatt_to_egsa_button.grid(                                  column=3, row=1                                )

        self.egsa_label.grid(                                           column=4, row=1                                )

        self.progress_text.set(" ")
        self.progress_label.grid(                                       column=2, row=2                                )

        self.main_frame.grid(                                                     row=1,            padx=10, pady=10   )

        self.quit_button["command"] = self.destroy_self
        self.quit_button.grid(                                          column=2, row=0                                )

        self.exit_frame.grid(                                                     row=2,            padx=10, pady=10   )

    def change_file_in(self):
        in_filename = filedialog.askopenfilename(initialdir=script_path, title="Select file to use ai Input")
        if os.path.exists(in_filename):
            self.in_filename = in_filename
            self.in_file_text.set(self.in_filename)

    def change_file_out(self):

        out_filename = filedialog.asksaveasfilename(initialdir=script_path, title="Select file")
        if out_filename is not "":
            self.out_filename = out_filename
            self.out_file_text.set(self.out_filename)

    def open_file_in_pros(self):
        os.startfile(self.in_filename)

    def open_file_out_pros(self):
        os.startfile(self.out_filename)

    def fl_to_hatt(self, driver, row):

        xpaths = [
            "//*[@id='a1']",
            "//*[@id='a2']",
            "//*[@id='a3']",
            "//*[@id='b1']",
            "//*[@id='b2']",
            "//*[@id='b3']"
        ]

        for i, item in enumerate(row):
            elem = driver.find_element_by_xpath(xpaths[i])
            elem.clear()
            elem.send_keys(str(0 if row[i] is None else row[i]))

        xpath = "/html/body/div[3]/div/ul/li[1]/div[3]/table/tbody/tr/td[1]/form/table/tbody/tr[10]/td/input"
        elem = driver.find_element_by_xpath(xpath)
        elem.click()

        elem = driver.find_element_by_xpath("//*[@id='r1']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        hatt_x = self.clipboard_get()

        elem = driver.find_element_by_xpath("//*[@id='r2']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        hatt_y = self.clipboard_get()

        return hatt_x, hatt_y

    def hatt_to_egsa(self, driver, row):

        elem = driver.find_element_by_xpath("//*[@id='a1']")
        elem.clear()
        elem.send_keys(str(0 if row[0] is None else row[0]))

        elem = driver.find_element_by_xpath("//*[@id='b1']")
        elem.clear()
        elem.send_keys(str(0 if row[1] is None else row[1]))

        driver.find_element_by_xpath("//*[@id='Μετατροπή2']").click()

        elem = driver.find_element_by_xpath("//*[@id='a2']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        egsa_x = self.clipboard_get()

        elem = driver.find_element_by_xpath("//*[@id='b2']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        egsa_y = self.clipboard_get()

        return egsa_x, egsa_y

    def start_driver(self):
        if self.driver_state == "stopped":
            self.driver = webdriver.Firefox(executable_path=r'geckodriver.exe')
            self.driver.minimize_window()
            self.driver_state = "running"

    def update_values(self):
        self.first_col = int(self.col_entry.get())
        self.first_row = int(self.row_entry.get())
        self.workingState = 0
        self.update_working_state()

    def update_working_state(self):
        if self.workingState == 0:
            self.progress_text.set("Working   ")
            self.workingState = 1
        elif self.workingState == 1:
            self.progress_text.set("Working.  ")
            self.workingState = 2
        elif self.workingState == 2:
            self.progress_text.set("Working.. ")
            self.workingState = 3
        elif self.workingState == 3:
            self.progress_text.set("Working...")
            self.workingState = 1

        self.update()

    def start_fl_to_hatt(self):
        if self.option.get() == "web":
            self.online_fl_to_hatt()
        elif self.option.get() == "no_web":
            self.offline_fl_to_hatt()

    def offline_fl_to_hatt(self):
        self.update_values()
        workbook = load_workbook(filename=self.in_filename)
        end_book = Workbook()
        first_sheet = end_book.active

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            end_book.create_sheet(sheet_name)
            out_sheet = end_book[sheet_name]

            out_sheet.cell(row=1, column=1).value = "id"
            out_sheet.cell(row=1, column=2).value = "x"
            out_sheet.cell(row=1, column=3).value = "y"

            for i, row in enumerate(sheet.iter_rows(min_row=self.first_row,
                                                    min_col=self.first_col,
                                                    max_col=self.first_col + 5,
                                                    values_only=True)):
                res_x = row[0] + (row[1]/60) + (row[2]/3600)
                res_y = row[3] + (row[4]/60) + (row[5]/3600)
                out_sheet.cell(row=i + 2, column=1).value = i + 1
                out_sheet.cell(row=i + 2, column=2).value = res_x
                out_sheet.cell(row=i + 2, column=3).value = res_y
                self.update_working_state()

        end_book.remove_sheet(first_sheet)
        self.change_file_out()
        end_book.save(self.out_filename)
        self.progress_text.set("Finished!")

    def online_fl_to_hatt(self):
        self.update_values()
        self.start_driver()
        workbook = load_workbook(filename=self.in_filename)
        end_book = Workbook()
        first_sheet = end_book.active

        url = "http://www.calcfun.com/calc-67-metatropi-syntetagmenon-apo-moires-lepta-deyterolepta-se-dekadikes.html"
        self.driver.get(url)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            end_book.create_sheet(sheet_name)
            out_sheet = end_book[sheet_name]

            out_sheet.cell(row=1, column=1).value = "id"
            out_sheet.cell(row=1, column=2).value = "x"
            out_sheet.cell(row=1, column=3).value = "y"

            for i, row in enumerate(sheet.iter_rows(min_row=self.first_row,
                                                    min_col=self.first_col,
                                                    max_col=self.first_col + 5,
                                                    values_only=True)):
                res = self.fl_to_hatt(self.driver, row)
                out_sheet.cell(row=i + 2, column=1).value = i + 1
                out_sheet.cell(row=i + 2, column=2).value = float(res[0])
                out_sheet.cell(row=i + 2, column=3).value = float(res[1])
                self.update_working_state()

        end_book.remove_sheet(first_sheet)
        self.change_file_out()
        end_book.save(self.out_filename)
        self.progress_text.set("Finished!")

    def start_hatt_to_egsa(self):
        if self.option.get() == "web":
            self.online_hatt_to_esga()
        elif self.option.get() == "no_web":
            self.offline_hatt_to_esga()

    def offline_hatt_to_esga(self):
        print("offline_hatt_to_esga" + self.option.get())

    def online_hatt_to_esga(self):
        self.update_values()
        self.start_driver()
        workbook = load_workbook(filename=self.in_filename)
        end_book = Workbook()
        first_sheet = end_book.active

        url = "http://www.calcfun.com/calc-75-metatropi-syntetagmenon-apo-wgs84-se-egsa-87.html"
        self.driver.get(url)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            end_book.create_sheet(sheet_name)
            out_sheet = end_book[sheet_name]

            out_sheet.cell(row=1, column=1).value = "id"
            out_sheet.cell(row=1, column=2).value = "x"
            out_sheet.cell(row=1, column=3).value = "y"

            for i, row in enumerate(sheet.iter_rows(min_row=self.first_row,
                                                    min_col=self.first_col,
                                                    max_col=self.first_col + 1,
                                                    values_only=True)):
                res = self.hatt_to_egsa(self.driver, row)

                out_sheet.cell(row=i + 2, column=1).value = i + 1
                out_sheet.cell(row=i + 2, column=2).value = float(res[0])
                out_sheet.cell(row=i + 2, column=3).value = float(res[1])
                self.update_working_state()

        end_book.remove_sheet(first_sheet)
        self.change_file_out()
        end_book.save(self.out_filename)
        self.progress_text.set("Finished!")


root = tk.Tk()
# root.protocol("WM_DELETE_WINDOW", root.maxsize)
app = Application(master=root)
root.protocol("WM_DELETE_WINDOW", app.destroy_self)
app.mainloop()
