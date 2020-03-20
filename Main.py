from openpyxl import load_workbook
import tkinter as tk
from tkinter.filedialog import askopenfilename
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import os


# noinspection DuplicatedCode
class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)

        # Global variables init ----------------------------------------------------------------------------------------
        self.in_filename = "coordinates.xlsx"
        self.out_filename = "coordinates.xlsx"
        self.driver = webdriver.Firefox(executable_path=r'geckodriver.exe')

        # Window init --------------------------------------------------------------------------------------------------
        self.master = master
        master.resizable(False, False)
        self.grid()

        # Window frames init -------------------------------------------------------------------------------------------
        self.file_frame = tk.Frame(self.master, relief="sunken", bg="grey")
        self.main_frame = tk.Frame(self.master)

        # Variable strings init ----------------------------------------------------------------------------------------
        self.in_file_text = tk.StringVar()
        self.out_file_text = tk.StringVar()
        self.progress_text = tk.StringVar()

        # Label init ---------------------------------------------------------------------------------------------------
        self.tip_label_in = tk.Label(self.file_frame,   text="Input file:   ",            bg="grey")
        self.tip_label_out = tk.Label(self.file_frame,  text="Output file:  ",            bg="grey")
        self.in_file_label = tk.Label(self.file_frame,  textvariable=self.in_file_text,   bg="grey")
        self.out_file_label = tk.Label(self.file_frame, textvariable=self.out_file_text,  bg="grey")
        self.fl_label = tk.Label(self.main_frame,       text="     Φ,Λ     ")
        self.hatt_label = tk.Label(self.main_frame,     text="        hatt        ")
        self.egsa_label = tk.Label(self.main_frame,     text="     egsa     ")
        self.progress_label = tk.Label(self.main_frame, textvariable=self.progress_text,  fg="green", padx=10, pady=10)

        # Button init --------------------------------------------------------------------------------------------------
        self.select_file_in_button = tk.Button(self.file_frame,     bg="grey")
        self.open_file_in_button = tk.Button(self.file_frame,       bg="grey")
        self.select_file_out_button = tk.Button(self.file_frame,    bg="grey")
        self.open_file_out_button = tk.Button(self.file_frame,      bg="grey")
        self.fl_to_hatt_button = tk.Button(self.main_frame)
        self.hatt_to_egsa_button = tk.Button(self.main_frame)
        self.quit_button = tk.Button(self.main_frame, text="QUIT",  fg="red")
        self.place_widgets()

    def _destroy(self):
        self.driver.quit()
        try:
            os.remove("geckodriver.log")
        except PermissionError:
            print("Exception raised: PermissionError. File: geckodriver.log")
        self.master.destroy()

    def place_widgets(self):

        # File frame widgets -------------------------------------------------------------------------------------------
        self.tip_label_in.grid(                                         column=0, row=0, sticky='w'                    )
        self.in_file_text.set(self.in_filename)
        self.in_file_label.grid(                                        column=1, row=0, sticky='w'                    )

        self.select_file_in_button["text"] = "Change input file...  "
        self.select_file_in_button["command"] = self.change_file_in
        self.select_file_in_button.grid(                                column=2, row=0,            padx=5, pady=5     )

        self.open_file_in_button["text"] = "Open input file.  "
        self.open_file_in_button["command"] = self.open_file_in_pros
        self.open_file_in_button.grid(                                  column=3, row=0,            padx=10            )

        self.tip_label_out.grid(                                        column=0, row=1, sticky='w'                    )
        self.out_file_text.set(self.out_filename)
        self.out_file_label.grid(                                       column=1, row=1, sticky='w'                    )

        self.select_file_out_button["text"] = "Change output file..."
        self.select_file_out_button["command"] = self.change_file_out
        self.select_file_out_button.grid(                               column=2, row=1,            padx=5             )

        self.open_file_out_button["text"] = "Open output file."
        self.open_file_out_button["command"] = self.open_file_out_pros
        self.open_file_out_button.grid(                                 column=3, row=1,            padx=10, pady=5    )

        self.file_frame.grid(                                                     row=0,            padx=10, pady=10   )

        # Main widgets -------------------------------------------------------------------------------------------------
        self.fl_label.grid(                                             column=0, row=0                                )

        self.fl_to_hatt_button["text"] = "     -->     "
        self.fl_to_hatt_button["command"] = self.start_fl_to_hatt
        self.fl_to_hatt_button.grid(                                    column=1, row=0                                )

        self.hatt_label.grid(                                           column=2, row=0                                )

        self.hatt_to_egsa_button["text"] = "     -->     "
        self.hatt_to_egsa_button["command"] = self.start_hatt_to_egsa
        self.hatt_to_egsa_button.grid(                                  column=3, row=0                                )

        self.egsa_label.grid(                                           column=4, row=0                                )

        self.progress_text.set(" ")
        self.progress_label.grid(                                       column=2, row=1                                )

        self.quit_button["command"] = self._destroy
        self.quit_button.grid(                                          column=2, row=2                                )

        self.main_frame.grid(                                                     row=1,            padx=10, pady=10   )

    def change_file_in(self):
        self.in_filename = askopenfilename()
        self.in_file_text.set("Input File is:    " + self.in_filename)

    def change_file_out(self):
        self.out_filename = askopenfilename()
        self.out_file_text.set("Output File is: " + self.out_filename)

    def open_file_in_pros(self):
        os.startfile(self.in_filename)

    def open_file_out_pros(self):
        os.startfile(self.out_filename)

    def fl_to_hatt(self, driver, row):

        xpaths = [
            "//*[@id='a1']",
            "//*[@id='a2']",
            "//*[@id='a3']",
            "//*[@id='b1']",
            "//*[@id='b2']",
            "//*[@id='b3']"
        ]

        for i, item in enumerate(row):
            elem = driver.find_element_by_xpath(xpaths[i])
            elem.clear()
            elem.send_keys(str(0 if row[i] is None else row[i]))

        xpath = "/html/body/div[3]/div/ul/li[1]/div[3]/table/tbody/tr/td[1]/form/table/tbody/tr[10]/td/input"
        elem = driver.find_element_by_xpath(xpath)
        elem.click()

        elem = driver.find_element_by_xpath("//*[@id='r1']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        hatt_x = self.clipboard_get()

        elem = driver.find_element_by_xpath("//*[@id='r2']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        hatt_y = self.clipboard_get()

        return hatt_x, hatt_y

    def hatt_to_egsa(self, driver, row):

        elem = driver.find_element_by_xpath("//*[@id='a1']")
        elem.clear()
        elem.send_keys(str(0 if row[0] is None else row[0]))

        elem = driver.find_element_by_xpath("//*[@id='b1']")
        elem.clear()
        elem.send_keys(str(0 if row[1] is None else row[1]))

        driver.find_element_by_xpath("//*[@id='Μετατροπή2']").click()

        elem = driver.find_element_by_xpath("//*[@id='a2']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        egsa_x = self.clipboard_get()

        elem = driver.find_element_by_xpath("//*[@id='b2']")
        elem.send_keys(Keys.CONTROL + "a")
        elem.send_keys(Keys.CONTROL + "c")
        egsa_y = self.clipboard_get()

        return egsa_x, egsa_y

    def start_fl_to_hatt(self):
        self.progress_text.set("Working...")
        self.update()
        workbook = load_workbook(filename=self.in_filename)

        url = "http://www.calcfun.com/calc-67-metatropi-syntetagmenon-apo-moires-lepta-deyterolepta-se-dekadikes.html"
        self.driver.get(url)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for i, row in enumerate(sheet.iter_rows(min_row=3, min_col=2, max_col=7, values_only=True)):

                res = self.fl_to_hatt(self.driver, row)

                sheet.cell(row=i + 3, column=8).value = res[0]
                sheet.cell(row=i + 3, column=9).value = res[1]

        workbook.save(self.out_filename)
        self.progress_text.set("Finished!")

    def start_hatt_to_egsa(self):
        self.progress_text.set("Working...")
        self.update()
        workbook = load_workbook(filename=self.in_filename)

        url = "http://www.calcfun.com/calc-75-metatropi-syntetagmenon-apo-wgs84-se-egsa-87.html"
        self.driver.get(url)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for i, row in enumerate(sheet.iter_rows(min_row=3, min_col=8, max_col=9, values_only=True)):

                res = self.hatt_to_egsa(self.driver, row)

                sheet.cell(row=i + 3, column=10).value = res[0]
                sheet.cell(row=i + 3, column=11).value = res[1]

        workbook.save(self.out_filename)
        self.progress_text.set("Finished!")


root = tk.Tk()
root.protocol("WM_DELETE_WINDOW", root.maxsize)
app = Application(master=root)
app.mainloop()
